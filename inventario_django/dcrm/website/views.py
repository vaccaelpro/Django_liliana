import re
import html
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.core.paginator import Paginator
from .models import Usuario, LogAuditoria

# ============================================================
# CAPA 3 — Sanitización de inputs (html.escape)
# ============================================================
def sanitizar(valor):
    """
    Capa 3 - Sanitización: elimina caracteres HTML peligrosos del input.
    Previene XSS al escapar <, >, &, ', " antes de procesar cualquier dato.
    """
    return html.escape(str(valor).strip()) if valor else ''


# --- PATRONES DE VALIDACIÓN ---
REGEX_DOCUMENTO = re.compile(r'^\d{5,12}$')
REGEX_NOMBRES   = re.compile(r'^[a-zA-ZáéíóúÁÉÍÓÚñÑ ]{2,50}$')
REGEX_EMAIL     = re.compile(r'^[\w\.-]+@[\w\.-]+\.\w{2,}$')
REGEX_PASSWORD  = re.compile(r'^(?=.*[A-Z])(?=.*\d).{8,}$')


# --- HELPERS ---
def es_administrador(user):
    return user.is_authenticated and user.rol == 'ADMINISTRADOR'

def es_aprendiz(user):
    return user.is_authenticated and user.rol == 'APRENDIZ'

def registrar_log(usuario, accion, detalle, request):
    """Registra acciones de auditoría con IP del cliente."""
    ip = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', ''))
    if ',' in ip:
        ip = ip.split(',')[0].strip()
    LogAuditoria.objects.create(
        usuario=usuario,
        accion=accion,
        detalle=detalle,
        ip=ip
    )

def get_ip(request):
    """Extrae la IP real del cliente, considerando proxies."""
    ip = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', 'unknown'))
    return ip.split(',')[0].strip() if ',' in ip else ip


# ============================================================
# VISTAS DE ACCESO
# ============================================================

def home(request):
    return redirect('login')


def login_user(request):
    if request.user.is_authenticated:
        if request.user.rol == 'ADMINISTRADOR':
            return redirect('admin_dashboard')
        return redirect('apprentice_dashboard')

    if request.method == 'POST':
        # Capa 3: sanitizar antes de validar
        documento = sanitizar(request.POST.get('documento', ''))
        password  = request.POST.get('password', '')

        if not REGEX_DOCUMENTO.match(documento):
            messages.error(request, "El documento debe contener entre 5 y 12 dígitos numéricos.")
            return render(request, 'login.html')

        if len(password) < 8:
            messages.error(request, "La contraseña debe tener al menos 8 caracteres.")
            return render(request, 'login.html')

        try:
            user_obj = Usuario.objects.get(documento=documento)
            user = authenticate(request, username=user_obj.username, password=password)

            if user is not None:
                login(request, user)
                registrar_log(user, 'LOGIN', f'Inicio de sesión: {user.documento}', request)
                messages.success(request, f"¡Bienvenido, {user.first_name or user.documento}!")
                if user.rol == 'ADMINISTRADOR':
                    return redirect('admin_dashboard')
                return redirect('apprentice_dashboard')
            else:
                messages.error(request, "Contraseña incorrecta.")
        except Usuario.DoesNotExist:
            messages.error(request, "El documento ingresado no está registrado.")

    return render(request, 'login.html')


@require_POST
def register_user(request):
    # Capa 3: sanitizar todos los campos de texto
    nombres   = sanitizar(request.POST.get('nombres', ''))
    apellidos = sanitizar(request.POST.get('apellidos', ''))
    email     = sanitizar(request.POST.get('email', ''))
    documento = sanitizar(request.POST.get('documento', ''))
    rol       = request.POST.get('rol', '')
    password  = request.POST.get('password', '')

    if not REGEX_NOMBRES.match(nombres):
        messages.error(request, "Los nombres solo deben contener letras y espacios (2–50 caracteres).")
        return redirect('login')
    if not REGEX_NOMBRES.match(apellidos):
        messages.error(request, "Los apellidos solo deben contener letras y espacios (2–50 caracteres).")
        return redirect('login')
    if not REGEX_EMAIL.match(email):
        messages.error(request, "El correo electrónico no tiene un formato válido.")
        return redirect('login')
    if not REGEX_DOCUMENTO.match(documento):
        messages.error(request, "El documento debe contener entre 5 y 12 dígitos numéricos.")
        return redirect('login')
    if not REGEX_PASSWORD.match(password):
        messages.error(request, "La contraseña debe tener mínimo 8 caracteres, una mayúscula y un número.")
        return redirect('login')
    if rol not in ('ADMINISTRADOR', 'APRENDIZ'):
        messages.error(request, "El rol seleccionado no es válido.")
        return redirect('login')
    if Usuario.objects.filter(documento=documento).exists():
        messages.error(request, "Ya existe un usuario con ese número de documento.")
        return redirect('login')
    if Usuario.objects.filter(email=email).exists():
        messages.error(request, "Ya existe una cuenta registrada con ese correo electrónico.")
        return redirect('login')

    nuevo_usuario = Usuario.objects.create_user(
        username=documento,
        password=password,
        first_name=nombres,
        last_name=apellidos,
        email=email,
        documento=documento,
        rol=rol,
    )
    registrar_log(nuevo_usuario, 'REGISTRO', f'Nuevo usuario registrado: {documento}', request)
    messages.success(request, "Cuenta creada exitosamente. Ahora puedes iniciar sesión.")
    return redirect('login')


def logout_user(request):
    if request.user.is_authenticated:
        registrar_log(request.user, 'LOGOUT', f'Cierre de sesión: {request.user.documento}', request)
    logout(request)
    messages.success(request, "Has cerrado sesión correctamente.")
    return redirect('login')


# ============================================================
# DASHBOARDS
# ============================================================

@login_required(login_url='login')
@user_passes_test(es_administrador, login_url='login')
def admin_dashboard(request):
    return render(request, 'admin_dashboard.html')


@login_required(login_url='login')
@user_passes_test(es_administrador, login_url='login')
def gestion_usuarios(request):
    users = Usuario.objects.all().order_by('date_joined')
    query = sanitizar(request.GET.get('q', ''))
    if query:
        users = users.filter(first_name__icontains=query) | users.filter(documento__icontains=query)

    paginator = Paginator(users, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'gestion_usuarios.html', {
        'users': page_obj,
        'page_obj': page_obj,
        'query': query,
    })


@login_required(login_url='login')
@user_passes_test(es_administrador, login_url='login')
def logs_auditoria(request):
    logs = LogAuditoria.objects.all().order_by('-fecha')[:50]
    return render(request, 'logs_auditoria.html', {
        'logs': logs,
    })


@login_required(login_url='login')
@user_passes_test(es_administrador, login_url='login')
def export_users_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    # Definir estilos
    header_font = Font(name='Calibri', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid") # SENA Green
    alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['Documento', 'Nombres', 'Apellidos', 'Correo', 'Rol', 'Fecha Registro']
    ws.append(headers)

    # Aplicar estilos a las cabeceras
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border
        # Ajustar ancho de columna aproximado
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25

    users = Usuario.objects.all().order_by('date_joined')
    for row_num, u in enumerate(users, 2):
        row = [
            u.documento, u.first_name, u.last_name, u.email, 
            u.get_rol_display(), u.date_joined.strftime("%Y-%m-%d %H:%M")
        ]
        ws.append(row)
        for col_num, cell in enumerate(ws[row_num], 1):
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="usuarios.xlsx"'
    wb.save(response)
    
    registrar_log(request.user, 'REPORTE', 'Exportó listado de usuarios a Excel', request)
    return response


@login_required(login_url='login')
@user_passes_test(es_administrador, login_url='login')
def export_logs_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Logs de Auditoría"

    # Definir estilos
    header_font = Font(name='Calibri', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
    alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['Fecha', 'Usuario', 'Acción', 'Detalle', 'IP']
    ws.append(headers)

    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25
    ws.column_dimensions['D'].width = 50 # Detalle más ancho

    logs = LogAuditoria.objects.all().order_by('-fecha')
    for row_num, log in enumerate(logs, 2):
        usuario_str = log.usuario.first_name if log.usuario else "—"
        row = [
            log.fecha.strftime("%Y-%m-%d %H:%M"), usuario_str, 
            log.get_accion_display(), log.detalle, log.ip
        ]
        ws.append(row)
        for col_num, cell in enumerate(ws[row_num], 1):
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="logs_auditoria.xlsx"'
    wb.save(response)
    
    registrar_log(request.user, 'REPORTE', 'Exportó logs de auditoría a Excel', request)
    return response


@login_required(login_url='login')
@user_passes_test(es_aprendiz, login_url='login')
def apprentice_dashboard(request):
    return render(request, 'apprentice_dashboard.html')


# ============================================================
# CRUD
# ============================================================

@login_required(login_url='login')
@user_passes_test(es_administrador, login_url='login')
@require_POST
def user_delete(request, user_id):
    if request.user.id == user_id:
        messages.error(request, "No puedes eliminar tu propia cuenta.")
        return redirect('admin_dashboard')

    user = get_object_or_404(Usuario, id=user_id)
    registrar_log(request.user, 'ELIMINACION', f'Usuario eliminado: {user.documento} - {user.first_name}', request)
    user.delete()
    messages.success(request, "Usuario eliminado correctamente.")
    return redirect('admin_dashboard')


@login_required(login_url='login')
@user_passes_test(es_administrador, login_url='login')
def edit_user(request, user_id):
    user_to_edit = get_object_or_404(Usuario, id=user_id)

    if request.method == 'POST':
        nombres      = sanitizar(request.POST.get('nombres', ''))
        apellidos    = sanitizar(request.POST.get('apellidos', ''))
        email        = sanitizar(request.POST.get('email', ''))
        documento    = sanitizar(request.POST.get('documento', ''))
        rol          = request.POST.get('rol', '')
        new_password = request.POST.get('password', '')

        errores = {}
        if not REGEX_NOMBRES.match(nombres):
            errores['nombres'] = "Solo letras y espacios (2–50 caracteres)."
        if not REGEX_NOMBRES.match(apellidos):
            errores['apellidos'] = "Solo letras y espacios (2–50 caracteres)."
        if not REGEX_EMAIL.match(email):
            errores['email'] = "Formato de correo no válido."
        if not REGEX_DOCUMENTO.match(documento):
            errores['documento'] = "Entre 5 y 12 dígitos numéricos."
        if new_password and not REGEX_PASSWORD.match(new_password):
            errores['password'] = "Mínimo 8 caracteres, una mayúscula y un número."
        if rol not in ('ADMINISTRADOR', 'APRENDIZ'):
            errores['rol'] = "Rol no válido."
        if Usuario.objects.filter(documento=documento).exclude(id=user_id).exists():
            errores['documento'] = "Ese documento ya está en uso por otro usuario."
        if Usuario.objects.filter(email=email).exclude(id=user_id).exists():
            errores['email'] = "Ese correo ya está en uso por otro usuario."

        if errores:
            for msg in errores.values():
                messages.error(request, msg)
            return render(request, 'edit_user.html', {
                'user_to_edit': user_to_edit,
                'es_admin': user_to_edit.rol == 'ADMINISTRADOR',
            })

        user_to_edit.first_name = nombres
        user_to_edit.last_name  = apellidos
        user_to_edit.email      = email
        user_to_edit.documento  = documento
        user_to_edit.rol        = rol

        if new_password:
            user_to_edit.set_password(new_password)

        registrar_log(request.user, 'EDICION', f'Editó usuario: {user_to_edit.documento} - {user_to_edit.first_name}', request)
        user_to_edit.save()
        messages.success(request, f"Datos de {user_to_edit.first_name} actualizados correctamente.")
        return redirect('admin_dashboard')

    return render(request, 'edit_user.html', {
        'user_to_edit': user_to_edit,
        'es_admin': user_to_edit.rol == 'ADMINISTRADOR',
    })

@login_required(login_url='login')
@user_passes_test(es_aprendiz, login_url='login')
def cambiar_password_aprendiz(request):
    if request.method == 'POST':
        password_actual  = request.POST.get('password_actual', '')
        password_nueva   = request.POST.get('password_nueva', '')
        password_confirma = request.POST.get('password_confirma', '')

        if not request.user.check_password(password_actual):
            messages.error(request, "La contraseña actual es incorrecta.")
            return redirect('apprentice_dashboard')

        if not REGEX_PASSWORD.match(password_nueva):
            messages.error(request, "La nueva contraseña debe tener mínimo 8 caracteres, una mayúscula y un número.")
            return redirect('apprentice_dashboard')

        if password_nueva != password_confirma:
            messages.error(request, "Las contraseñas nuevas no coinciden.")
            return redirect('apprentice_dashboard')

        request.user.set_password(password_nueva)
        request.user.save()
        registrar_log(request.user, 'CAMBIO_PASSWORD', 'Aprendiz cambió su contraseña', request)
        messages.success(request, "Contraseña actualizada correctamente. Inicia sesión de nuevo.")
        return redirect('login')

    return redirect('apprentice_dashboard')

@login_required(login_url='login')
@user_passes_test(es_aprendiz, login_url='login')
def perfil_aprendiz(request):
    if request.method == 'POST':
        nombres   = sanitizar(request.POST.get('nombres', ''))
        apellidos = sanitizar(request.POST.get('apellidos', ''))
        email     = sanitizar(request.POST.get('email', ''))

        errores = {}
        if not REGEX_NOMBRES.match(nombres):
            errores['nombres'] = "Solo letras y espacios (2–50 caracteres)."
        if not REGEX_NOMBRES.match(apellidos):
            errores['apellidos'] = "Solo letras y espacios (2–50 caracteres)."
        if not REGEX_EMAIL.match(email):
            errores['email'] = "Formato de correo no válido."
        if Usuario.objects.filter(email=email).exclude(id=request.user.id).exists():
            errores['email'] = "Ese correo ya está en uso por otro usuario."

        if errores:
            for msg in errores.values():
                messages.error(request, msg)
            return render(request, 'perfil_aprendiz.html')

        request.user.first_name = nombres
        request.user.last_name  = apellidos
        request.user.email      = email
        request.user.save()
        registrar_log(request.user, 'EDICION', 'Aprendiz actualizó su perfil', request)
        messages.success(request, "Perfil actualizado correctamente.")
        return redirect('perfil_aprendiz')

    return render(request, 'perfil_aprendiz.html')


# ============================================================
# ERRORES PERSONALIZADOS
# ============================================================

def error_404(request, exception):
    return render(request, '404.html', status=404)

def error_500(request):
    return render(request, '500.html', status=500)